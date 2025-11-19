import openpyxl as excel

#엑셀 파일을 불러옴
book = excel.load_workbook("py_excel01.xlsx")
sheet = book.worksheets[0]
#sheet = book.active

sheet.cell(row=2, column=1, value="행열을 이용한 파이썬 엑셀 사례")

row_cell = sheet.cell(row=3, column=1)
row_cell.value = "행영을 이용한 파이썬 두번째 사례"

book.save("py_excel02.xlsx")